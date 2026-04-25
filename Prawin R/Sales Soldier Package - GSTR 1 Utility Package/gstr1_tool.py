import glob
import json
import os
import re
import shutil
import subprocess
import tempfile
from datetime import datetime

import pandas as pd
import tkinter as tk
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import filedialog, messagebox, scrolledtext, ttk


APP_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(APP_DIR, "master.xlsx")
FINANCIAL_YEARS = ["2025-26", "2026-27", "2027-28", "2028-29", "2029-30"]

INTRODUCTION_TEXT = """Welcome to Sales Soldier

Sales Soldier is a smart and practical GST support application designed especially for small, micro, and medium taxpayers who find it difficult to prepare and file GSTR-1 returns on their own.

Many small business owners have limited sales volume and cannot always afford professional compliance support every month. At the same time, filing GSTR-1 manually through the GST Portal or through the Offline Utility can be confusing, time-consuming, and technically difficult.
Sales Soldier was created to solve that exact problem.

Why Sales Soldier Was Created

The idea for Sales Soldier came from observing real taxpayers who struggled with:

Segregating B2B and B2C sales
- Identifying Inter-State and Intra-State supplies
- Preparing HSN-wise summaries
- Maintaining proper invoice/document numbering
- Understanding the GST Portal
- Using the GSTR-1 Offline Utility
- Filling Excel / CSV templates correctly
- Generating files for upload without errors

Sales Soldier simplifies all these tasks into one easy software.

🛡️ Mission of Sales Soldier:
To become a digital soldier for taxpayers, helping them fight complexity, save money, and stay GST compliant with confidence.

👨‍💻 Developed By

Prawin Ramesh Kannan
(CA Finalist)

🙏 Credits

Special thanks to Rathinagiri Sir.

Thanks to my friends for their valuable suggestions.
"""

HELP_TEXT = """Step 1: Basic Information
1. CLIENT: Enter the name of your business / company.
2. GSTN: Enter your GST registration number. If unregistered, type URP.
3. FY: Select the relevant Financial Year.
4. MONTH: Select the month for which invoices belong
(Example: March 2026).

Step 2: Invoice Details

This section captures each invoice.
1. INVOICE NUMBER: Enter invoice number.
2. INVOICE DATE: Enter invoice date.
3. PARTY GSTN: Enter buyer GSTN.
4. SUPPLY TYPE
- B2B -> Automatically selected for registered party
- B2C-1 -> Intra-State B2C sales
- B2C-2 -> Inter-State B2C sales
5. PARTY NAME: Enter buyer / customer name.
6. PLACE OF SUPPLY: Enabled when B2C-2 is selected.

7. +ADD PARTY: Save party details into Party Master for future use.

8. CANCEL: For cancelled invoice:
- Fill Invoice Number
- Fill Invoice Date
- Click Cancel
- Other fields may remain disabled.

Step 3: Product / Service Section
1. PRODUCT / SERVICE: Enter item name.
2. GOODS / SERVICE: Select whether Goods or Service.
3. HSN / SAC: Enter HSN (Goods) or SAC (Service).
4. QUANTITY: Enter quantity sold.
5. UNIT / UQC: Choose unit such as:
- NOS
- KGS
- LTR
- MTR
- PCS

6. ITEM TAXABLE VALUE: Enter taxable amount of item.
7. GSTR RATE %: Enter applicable GST rate.
8. TAX TYPE: Automatically selected based on Inter-State / Intra-State supply.
9. + PRODUCT MASTER: Save item details for future reuse.
10. + ADD ITEM: Adds item to Item List.
Multiple items can be added in one invoice.
11. DELETE ITEM: Removes selected item.

Step 4: Tax Section
TAX CALCULATION
Tax fields auto-calculate:
IGST
or
CGST + SGST
Manual editing is disabled.

Step 5: Invoice Management
1. + ADD INVOICE: Saves invoice and moves to next invoice entry.
2. DELETE INVOICE: Deletes selected invoice.
3. CLEAR FORM: Clears all current entry fields.

Step 6: Draft / Import Features
1. LOAD DRAFT: Loads previously saved draft data.
Example:
1 March -> saved invoices 1 to 5
2 March -> continue from same draft

2. IMPORT EXCEL: Imports invoice data from Excel.

Step 7: Export Features
1. EXPORT EXCEL: Exports overall monthly invoice data.
2. EXPORT GSTR-1 EXCEL: Exports overall monthly invoice data. Exports GST-ready structured file.
3. TRANSFER:
Transfers exported GSTR-1 Excel data into Offline Utility template.

Step 8: GST Portal Filing
GST LOGIN (Under Development)
Future feature for direct GST Portal login.

Final Filing Process
- Export GSTR-1 Excel
- Transfer to Offline Utility template
- Import in GST Offline Utility
- Generate JSON file
- Upload JSON to GST Portal
- File GSTR-1 Return

Important Notes
- Verify GSTN before filing
- Ensure invoice numbers are unique
- Review tax rates before export
- Keep backup of monthly files
"""

PARTY_COLUMNS = ["GSTN", "Party Name", "State"]
PRODUCT_COLUMNS = ["Product", "Item Type", "HSN", "Unit", "Rate"]
ITEM_TYPES = ["Goods", "Service"]
UQC_OPTIONS = [
    "",
    "BAG",
    "BOX",
    "BTL",
    "DOZ",
    "GMS",
    "KGS",
    "LTR",
    "MTR",
    "NOS",
    "PAC",
    "PCS",
    "QTL",
    "SET",
]
UQC_DISPLAY = {
    "BAG": "BAG-BAGS",
    "BOX": "BOX-BOX",
    "BTL": "BTL-BOTTLES",
    "DOZ": "DOZ-DOZENS",
    "GMS": "GMS-GRAMS",
    "KGS": "KGS-KILOGRAMS",
    "LTR": "LTR-LITRES",
    "MTR": "MTR-METERS",
    "NOS": "NOS-NUMBERS",
    "PAC": "PAC-PACKS",
    "PCS": "PCS-PIECES",
    "QTL": "QTL-QUINTAL",
    "SET": "SET-SETS",
}
SUPPLY_TYPES_REGISTERED = ["B2B"]
SUPPLY_TYPES_URP = ["B2C-1", "B2C-2"]
STATE_OPTIONS = [
    "01 - Jammu & Kashmir",
    "02 - Himachal Pradesh",
    "03 - Punjab",
    "04 - Chandigarh",
    "05 - Uttarakhand",
    "06 - Haryana",
    "07 - Delhi",
    "08 - Rajasthan",
    "09 - Uttar Pradesh",
    "10 - Bihar",
    "11 - Sikkim",
    "12 - Arunachal Pradesh",
    "13 - Nagaland",
    "14 - Manipur",
    "15 - Mizoram",
    "16 - Tripura",
    "17 - Meghalaya",
    "18 - Assam",
    "19 - West Bengal",
    "20 - Jharkhand",
    "21 - Odisha",
    "22 - Chhattisgarh",
    "23 - Madhya Pradesh",
    "24 - Gujarat",
    "25 - Daman & Diu",
    "26 - Dadra & Nagar Haveli",
    "27 - Maharashtra",
    "28 - Andhra Pradesh",
    "29 - Karnataka",
    "30 - Goa",
    "31 - Lakshadweep",
    "32 - Kerala",
    "33 - Tamil Nadu",
    "34 - Puducherry",
    "35 - Andaman & Nicobar Islands",
    "36 - Telangana",
    "37 - Andhra Pradesh",
    "38 - Ladakh",
    "97 - Other Territory",
]

INVOICE_COLUMNS = [
    "Financial Year",
    "Month",
    "Invoice Status",
    "Client Name",
    "Client GSTN",
    "Invoice No",
    "Invoice Date",
    "Party GSTN",
    "Party Name",
    "Party State Code",
    "Place of Supply",
    "Supply Type",
    "Tax Type",
    "Product/Service",
    "Item Type",
    "HSN",
    "Quantity",
    "Unit",
    "Taxable Value",
    "GST Rate",
    "IGST",
    "CGST",
    "SGST",
    "Total Invoice Value",
]


parties_df = pd.DataFrame(columns=PARTY_COLUMNS)
products_df = pd.DataFrame(columns=PRODUCT_COLUMNS)
invoices = []
editing_index = None
duplicate_invoice_warned = False
current_items = []
editing_item_index = None


def normalize_text(value):
    return str(value).strip()


def normalize_gstn(value):
    return normalize_text(value).upper()


def clean_month(value):
    value = normalize_text(value)
    value = re.sub(r"[^\w\- ]+", "", value)
    value = value.replace(" ", "_")
    return value


def get_month_options(financial_year):
    financial_year = normalize_text(financial_year)
    if not financial_year or "-" not in financial_year:
        return []
    start_year = int(financial_year.split("-")[0])
    months = []
    for month_number in range(4, 13):
        months.append(datetime(start_year, month_number, 1).strftime("%B-%Y").upper())
    for month_number in range(1, 4):
        months.append(datetime(start_year + 1, month_number, 1).strftime("%B-%Y").upper())
    return months


def parse_invoice_date(value):
    value = normalize_text(value)
    if not value:
        return ""

    compact = re.sub(r"\D", "", value)
    candidates = []
    if len(compact) == 8:
        candidates.append((compact, "%d%m%Y"))

    candidates.extend(
        [
            (value, "%d-%m-%Y"),
            (value, "%d/%m/%Y"),
            (value, "%d.%m.%Y"),
            (value, "%d-%m-%y"),
            (value, "%d/%m/%y"),
            (value, "%d-%b-%Y"),
            (value, "%d-%B-%Y"),
            (value, "%d %b %Y"),
            (value, "%d %B %Y"),
            (value, "%d-%b-%y"),
            (value, "%d-%B-%y"),
            (value, "%d %b %y"),
            (value, "%d %B %y"),
        ]
    )

    for candidate, date_format in candidates:
        try:
            return datetime.strptime(candidate.title(), date_format).strftime("%d-%b-%Y")
        except ValueError:
            continue
    raise ValueError("Invoice date must be in dd-mmm-yyyy format, for example 01-Mar-2026")


def validate_invoice_date_in_month(invoice_date, selected_month):
    if not invoice_date or not selected_month:
        return
    invoice_dt = datetime.strptime(invoice_date, "%d-%b-%Y")
    selected_dt = datetime.strptime(selected_month.title(), "%B-%Y")
    if invoice_dt.month != selected_dt.month or invoice_dt.year != selected_dt.year:
        raise ValueError("Enter the Correct date")


def parse_month(value):
    value = normalize_text(value)
    if not value:
        return ""

    compact = re.sub(r"[\s_\-/\.]", "", value).lower()
    candidates = [
        (compact, "%b%y"),
        (compact, "%B%y"),
        (compact, "%m%y"),
        (compact, "%b%Y"),
        (compact, "%B%Y"),
        (compact, "%m%Y"),
        (value.title(), "%b %y"),
        (value.title(), "%B %y"),
        (value.title(), "%m-%y"),
        (value.title(), "%m/%y"),
        (value.title(), "%b-%y"),
        (value.title(), "%B-%y"),
        (value.title(), "%b %Y"),
        (value.title(), "%B %Y"),
        (value.title(), "%m-%Y"),
        (value.title(), "%m/%Y"),
        (value.title(), "%b-%Y"),
        (value.title(), "%B-%Y"),
    ]

    for candidate, month_format in candidates:
        try:
            return datetime.strptime(candidate, month_format).strftime("%B-%Y").upper()
        except ValueError:
            continue
    raise ValueError("Month must include month and year, for example Mar26, March 26, or 03-26")


def to_float(value, field_name):
    try:
        return float(normalize_text(value))
    except ValueError:
        raise ValueError(f"{field_name} must be a number")


def same_text(series, value):
    value = normalize_text(value).lower()
    return series.fillna("").astype(str).str.strip().str.lower() == value


def contains_text(series, value):
    value = normalize_text(value).lower()
    if not value:
        return pd.Series([True] * len(series), index=series.index)
    return series.fillna("").astype(str).str.lower().str.contains(re.escape(value), na=False)


def natural_sort_key(value):
    value = normalize_text(value)
    return [
        int(part) if part.isdigit() else part.lower()
        for part in re.split(r"(\d+)", value)
        if part != ""
    ]


def natural_sort_tuple(value):
    return tuple(natural_sort_key(value))


def validate_gstn(gstn, field_name):
    gstn = normalize_gstn(gstn)
    if len(gstn) != 15:
        raise ValueError(f"{field_name} must be 15 characters")
    if not gstn[:2].isdigit():
        raise ValueError(f"{field_name} must start with a 2 digit state code")
    return gstn


def validate_party_gstn_or_urp(value):
    value = normalize_gstn(value)
    if value == "URP":
        return value
    return validate_gstn(value, "Party GSTN")


def state_code_from_option(value):
    value = normalize_text(value)
    if re.fullmatch(r"\d{2}", value):
        return value
    if " - " in value:
        return value.split(" - ", 1)[0]
    return ""


def state_option_from_code(code):
    code = normalize_text(code)
    for option in STATE_OPTIONS:
        if option.startswith(f"{code} - "):
            return option
    return ""


def ensure_master_file():
    if not os.path.exists(MASTER_FILE):
        with pd.ExcelWriter(MASTER_FILE, engine="openpyxl") as writer:
            pd.DataFrame(columns=PARTY_COLUMNS).to_excel(
                writer, sheet_name="Parties", index=False
            )
            pd.DataFrame(columns=PRODUCT_COLUMNS).to_excel(
                writer, sheet_name="Products", index=False
            )


def normalize_dataframe(df, columns):
    for column in columns:
        if column not in df.columns:
            df[column] = ""
    df = df[columns].copy()
    return df.fillna("")


def load_master():
    ensure_master_file()
    try:
        parties = pd.read_excel(MASTER_FILE, "Parties", dtype=str)
    except Exception:
        parties = pd.DataFrame(columns=PARTY_COLUMNS)

    try:
        products = pd.read_excel(MASTER_FILE, "Products", dtype=str)
    except Exception:
        products = pd.DataFrame(columns=PRODUCT_COLUMNS)

    parties = normalize_dataframe(parties, PARTY_COLUMNS)
    products = normalize_dataframe(products, PRODUCT_COLUMNS)
    parties["GSTN"] = parties["GSTN"].apply(normalize_gstn)
    products["Product"] = products["Product"].apply(normalize_text)
    products["Item Type"] = products["Item Type"].apply(normalize_text)
    products.loc[~products["Item Type"].isin(ITEM_TYPES), "Item Type"] = "Goods"
    products["Unit"] = products["Unit"].apply(lambda value: normalize_text(value).upper())
    return parties, products


def save_master():
    with pd.ExcelWriter(MASTER_FILE, engine="openpyxl", mode="w") as writer:
        parties_df.to_excel(writer, sheet_name="Parties", index=False)
        products_df.to_excel(writer, sheet_name="Products", index=False)


def get_draft_filename(month=None):
    try:
        month = clean_month(parse_month(month or month_var.get()))
    except ValueError:
        month = clean_month(month or month_var.get())
    if not month:
        return ""
    return os.path.join(APP_DIR, f"{month}_GSTR-1_draft.xlsx")


def get_export_filename():
    month = clean_month(parse_month(month_var.get()))
    if not month:
        raise ValueError("Enter month before export")

    base = f"{month}_GSTR-1"
    files = glob.glob(os.path.join(APP_DIR, f"{base}*.xlsx"))
    final_files = [
        file
        for file in files
        if not os.path.basename(file).lower().endswith("_draft.xlsx")
    ]

    used_versions = set()
    for file in final_files:
        name = os.path.splitext(os.path.basename(file))[0]
        if name == base:
            used_versions.add(0)
            continue
        match = re.fullmatch(rf"{re.escape(base)}_(\d+)", name)
        if match:
            used_versions.add(int(match.group(1)))

    version = 0
    while version in used_versions:
        version += 1

    if version == 0:
        return os.path.join(APP_DIR, f"{base}.xlsx")
    return os.path.join(APP_DIR, f"{base}_{version}.xlsx")


def get_final_gstr1_filename():
    month = clean_month(parse_month(month_var.get()))
    if not month:
        raise ValueError("Enter month before export")

    base = f"{month}_FINAL_GSTR-1"
    files = glob.glob(os.path.join(APP_DIR, f"{base}*.xlsx"))

    used_versions = set()
    for file in files:
        name = os.path.splitext(os.path.basename(file))[0]
        if name == base:
            used_versions.add(0)
            continue
        match = re.fullmatch(rf"{re.escape(base)}_(\d+)", name)
        if match:
            used_versions.add(int(match.group(1)))

    version = 0
    while version in used_versions:
        version += 1

    if version == 0:
        return os.path.join(APP_DIR, f"{base}.xlsx")
    return os.path.join(APP_DIR, f"{base}_{version}.xlsx")


def get_latest_final_gstr1_file():
    month = clean_month(parse_month(month_var.get()))
    if not month:
        raise ValueError("Enter month before transfer")

    pattern = os.path.join(APP_DIR, f"{month}_FINAL_GSTR-1*.xlsx")
    matches = [
        file
        for file in glob.glob(pattern)
        if os.path.isfile(file) and not os.path.basename(file).startswith("~")
    ]
    if not matches:
        raise FileNotFoundError(f"{month}_FINAL_GSTR-1 file not available")
    return max(matches, key=os.path.getmtime)


def get_template_source_file():
    for candidate in ["GSTR1_Template.xlsx", "GSTR1_Template.xlsm"]:
        candidate_path = os.path.join(APP_DIR, candidate)
        if os.path.exists(candidate_path):
            return candidate_path
    raise FileNotFoundError("GSTR1_Template not available")


def get_transfer_template_filename():
    month = clean_month(parse_month(month_var.get()))
    if not month:
        raise ValueError("Enter month before transfer")
    return os.path.join(APP_DIR, f"{month}_GSTR-1_TEMPLATE.xlsx")


def save_draft(show_error=False):
    file = get_draft_filename()
    if not file:
        return

    try:
        pd.DataFrame(get_invoice_summary_rows(), columns=INVOICE_COLUMNS).to_excel(file, index=False)
        status_var.set(f"Draft saved: {os.path.abspath(file)}")
    except Exception as exc:
        if show_error:
            messagebox.showerror("Draft Error", str(exc))


def load_draft():
    global invoices, editing_index
    editing_index = None
    file = get_draft_filename()
    invoices = []
    refresh_tree()

    if not file or not os.path.exists(file):
        status_var.set("No draft found for this month")
        update_button_state()
        return

    try:
        df = pd.read_excel(file, dtype=str).fillna("")
        df = normalize_dataframe(df, INVOICE_COLUMNS)
        loaded = normalize_dataframe(df, INVOICE_COLUMNS).to_dict("records")
        invoices = []
        for row in loaded:
            row["Items"] = []
            if normalize_text(row.get("Product/Service", "")):
                row["Items"] = [
                    {
                        "Product/Service": row.get("Product/Service", ""),
                        "Item Type": row.get("Item Type", ""),
                        "HSN": row.get("HSN", ""),
                        "Quantity": row.get("Quantity", ""),
                        "Unit": row.get("Unit", ""),
                        "Taxable Value": row.get("Taxable Value", ""),
                        "GST Rate": row.get("GST Rate", ""),
                        "IGST": row.get("IGST", ""),
                        "CGST": row.get("CGST", ""),
                        "SGST": row.get("SGST", ""),
                        "Total Value": row.get("Total Invoice Value", ""),
                    }
                ]
            invoices.append(row)
        refresh_tree()
        status_var.set(f"Loaded draft: {file}")
    except Exception as exc:
        messagebox.showerror("Load Error", str(exc))


def import_invoice_excel():
    global invoices, editing_index

    file = filedialog.askopenfilename(
        title="Select previous invoice Excel data",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
    )
    if not file:
        return

    try:
        excel_file = pd.ExcelFile(file)
        sheet_name = None
        for candidate in ["MAIN SHEET", "Invoices"]:
            if candidate in excel_file.sheet_names:
                sheet_name = candidate
                break
        if sheet_name is None:
            sheet_name = excel_file.sheet_names[0]

        df = pd.read_excel(file, sheet_name=sheet_name, dtype=str).fillna("")
        missing_columns = [
            column for column in INVOICE_COLUMNS if column not in df.columns
        ]
        if missing_columns:
            raise ValueError(
                "Selected file is not compatible. Missing columns: "
                + ", ".join(missing_columns)
            )

        imported_rows = normalize_dataframe(df, INVOICE_COLUMNS).to_dict("records")
        imported_invoices = []
        grouped = {}
        for row in imported_rows:
            key = normalize_text(row.get("Invoice No", ""))
            if key not in grouped:
                grouped[key] = {column: row.get(column, "") for column in INVOICE_COLUMNS}
                grouped[key]["Items"] = []
            if normalize_text(row.get("Product/Service", "")):
                grouped[key]["Items"].append(
                    {
                        "Product/Service": row.get("Product/Service", ""),
                        "Item Type": row.get("Item Type", ""),
                        "HSN": row.get("HSN", ""),
                        "Quantity": row.get("Quantity", ""),
                        "Unit": row.get("Unit", ""),
                        "Taxable Value": row.get("Taxable Value", ""),
                        "GST Rate": row.get("GST Rate", ""),
                        "IGST": row.get("IGST", ""),
                        "CGST": row.get("CGST", ""),
                        "SGST": row.get("SGST", ""),
                        "Total Value": row.get("Total Invoice Value", ""),
                    }
                )
        imported_invoices = list(grouped.values())
        if not imported_invoices:
            messagebox.showerror("Import Error", "Selected Excel file has no invoice rows")
            return

        if invoices:
            confirm = messagebox.askyesno(
                "Replace Current Data",
                "This will replace the current invoice list with the selected Excel data. Continue?",
            )
            if not confirm:
                return

        invoices = imported_invoices
        editing_index = None
        refresh_tree()
        clear_form(keep_header=True)
        save_draft()
        status_var.set(f"Imported {len(invoices)} invoices from {os.path.basename(file)}")
        messagebox.showinfo(
            "Import Complete",
            f"Imported {len(invoices)} invoices from:\n{file}",
        )
    except Exception as exc:
        messagebox.showerror("Import Error", str(exc))


def refresh_product_values():
    product_combo["values"] = get_product_options()


def get_party_gstn_options(filter_text=""):
    mask = contains_text(parties_df["GSTN"], filter_text)
    return sorted(parties_df.loc[mask, "GSTN"].dropna().astype(str).unique())


def get_party_name_options(filter_text=""):
    mask = contains_text(parties_df["Party Name"], filter_text)
    return sorted(parties_df.loc[mask, "Party Name"].dropna().astype(str).unique())


def get_product_options(filter_text=""):
    mask = contains_text(products_df["Product"], filter_text)
    return sorted(products_df.loc[mask, "Product"].dropna().astype(str).unique())


def open_combo_dropdown(combo):
    if combo["values"]:
        combo.event_generate("<Down>")


def filter_party_name_options(event=None):
    typed_text = party_name_var.get()
    party_name_combo["values"] = get_party_name_options(typed_text)
    open_combo_dropdown(party_name_combo)
    party_name_var.set(typed_text)
    party_name_combo.icursor(tk.END)


def filter_party_gstn_options(event=None):
    typed_text = party_gstn_var.get()
    party_gstn_combo["values"] = get_party_gstn_options(party_gstn_var.get())
    update_supply_type_controls()
    calculate_tax()
    open_combo_dropdown(party_gstn_combo)
    party_gstn_var.set(typed_text)
    party_gstn_combo.icursor(tk.END)


def filter_product_options(event=None):
    typed_text = product_var.get()
    product_combo["values"] = get_product_options(typed_text)
    open_combo_dropdown(product_combo)
    product_var.set(typed_text)
    product_combo.icursor(tk.END)


def defer_filter_party_gstn(event=None):
    root.after(1, filter_party_gstn_options)


def defer_filter_party_name(event=None):
    root.after(1, filter_party_name_options)


def defer_filter_product(event=None):
    root.after(1, filter_product_options)


def refresh_tree():
    for widget in invoice_strip_inner.winfo_children():
        widget.destroy()

    tk.Label(invoice_strip_inner, text="S.No", bg="#f2f2f2", font=("Arial", 10, "bold")).grid(
        row=0, column=0, padx=(6, 8), pady=4, sticky="w"
    )
    tk.Label(invoice_strip_inner, text="Inv.No", bg="#f2f2f2", font=("Arial", 10, "bold")).grid(
        row=1, column=0, padx=(6, 8), pady=4, sticky="w"
    )

    for index, invoice in enumerate(invoices, start=1):
        tk.Label(
            invoice_strip_inner,
            text=str(index),
            bg="#f2f2f2",
            width=8,
            anchor="center",
        ).grid(row=0, column=index, padx=4, pady=4)
        button = tk.Button(
            invoice_strip_inner,
            text=invoice.get("Invoice No", ""),
            width=8,
            command=lambda idx=index - 1: on_select(idx),
            bg="#3f5f8f" if editing_index == index - 1 else "#ffffff",
            fg="white" if editing_index == index - 1 else "#222222",
        )
        button.grid(row=1, column=index, padx=4, pady=4)

    invoice_strip_inner.update_idletasks()
    invoice_canvas.configure(scrollregion=invoice_canvas.bbox("all"))
    update_button_state()


def update_button_state():
    if editing_index is None:
        save_invoice_button.config(text="+ Add Invoice")
        delete_invoice_button.config(state="disabled")
    else:
        save_invoice_button.config(text="Update Invoice")
        delete_invoice_button.config(state="normal")


def clear_party_fields(event=None):
    party_gstn_var.set("")
    party_name_var.set("")
    party_state_var.set("")
    supply_type_var.set("B2B")
    tax_type_var.set("")
    update_supply_type_controls()
    calculate_tax()


def clear_party_because_gstn_empty(event=None):
    party_name_var.set("")
    party_state_var.set("")
    supply_type_var.set("B2B")
    tax_type_var.set("")
    update_supply_type_controls()
    calculate_tax()


def clear_party_because_name_empty(event=None):
    party_gstn_var.set("")
    party_state_var.set("")
    supply_type_var.set("B2B")
    tax_type_var.set("")
    update_supply_type_controls()
    calculate_tax()


def select_party_from_master_by_name(event=None):
    lookup_value = normalize_text(party_name_var.get())
    lookup_column = "Party Name"
    if not lookup_value:
        clear_party_fields()
        return

    match = parties_df[same_text(parties_df[lookup_column], lookup_value)]
    if not match.empty:
        party_gstn_var.set(match.iloc[0]["GSTN"])
        party_name_var.set(match.iloc[0]["Party Name"])
        update_supply_type_controls()
        calculate_tax()


def select_party_from_master_by_gstn(event=None):
    lookup_value = normalize_gstn(party_gstn_var.get())
    if not lookup_value:
        clear_party_fields()
        return
    if lookup_value == "URP":
        party_name_var.set("URP")
        update_supply_type_controls()
        calculate_tax()
        return

    match = parties_df[parties_df["GSTN"] == lookup_value]
    if not match.empty:
        party_gstn_var.set(match.iloc[0]["GSTN"])
        party_name_var.set(match.iloc[0]["Party Name"])
        update_supply_type_controls()
        calculate_tax()


def add_or_update_party():
    global parties_df
    try:
        gstn = validate_gstn(party_gstn_var.get(), "Party GSTN")
    except ValueError as exc:
        messagebox.showerror("Invalid GSTN", str(exc))
        return

    name = normalize_text(party_name_var.get())
    if not name:
        messagebox.showerror("Missing Data", "Enter party name")
        return

    if not parties_df[parties_df["GSTN"] == gstn].empty:
        messagebox.showinfo("Already Added", "This GSTN is already added in party master")
        return
    if not parties_df[same_text(parties_df["Party Name"], name)].empty:
        messagebox.showinfo("Already Added", "This party name is already added in party master")
        return

    row = {"GSTN": gstn, "Party Name": name, "State": gstn[:2]}
    parties_df = pd.concat([parties_df, pd.DataFrame([row])], ignore_index=True)
    messagebox.showinfo("Added", "Party added to master")

    save_master()
    party_gstn_combo["values"] = get_party_gstn_options()
    party_name_combo["values"] = get_party_name_options()


def clear_item_fields(event=None):
    product_var.set("")
    item_type_var.set("Goods")
    hsn_var.set("")
    quantity_var.set("1")
    unit_var.set("")
    rate_var.set("")
    update_quantity_controls()
    calculate_tax()


def clear_item_details(keep_product=False, keep_hsn=False):
    if not keep_product:
        product_var.set("")
    if not keep_hsn:
        hsn_var.set("")
    item_type_var.set("Goods")
    quantity_var.set("1")
    unit_var.set("")
    rate_var.set("")
    update_quantity_controls()
    calculate_tax()


def clear_item_because_product_empty(event=None):
    clear_item_details(keep_product=True)


def on_hsn_keyrelease(event=None):
    return


def select_product_from_master(event=None):
    product = normalize_text(product_var.get())
    if not product:
        clear_item_fields()
        return

    match = products_df[same_text(products_df["Product"], product)]
    if not match.empty:
        product_var.set(match.iloc[0]["Product"])
        item_type_var.set(match.iloc[0].get("Item Type", "Goods") or "Goods")
        hsn_var.set(match.iloc[0]["HSN"])
        unit_var.set(match.iloc[0].get("Unit", ""))
        rate_var.set(match.iloc[0]["Rate"])
        update_quantity_controls()
        calculate_tax()


def add_or_update_product():
    global products_df
    product = normalize_text(product_var.get())
    item_type = normalize_text(item_type_var.get()) or "Goods"
    hsn = normalize_text(hsn_var.get())
    unit = normalize_text(unit_var.get()).upper()

    if not product or not hsn:
        messagebox.showerror("Missing Data", "Enter product/service and HSN")
        return
    if item_type not in ITEM_TYPES:
        messagebox.showerror("Missing Data", "Select Goods or Service")
        return
    if item_type == "Goods" and not unit:
        messagebox.showerror("Missing Data", "Select quantity unit for goods")
        return

    try:
        rate = to_float(rate_var.get(), "GST rate")
    except ValueError as exc:
        messagebox.showerror("Invalid Rate", str(exc))
        return

    if item_type == "Service":
        unit = ""

    if not products_df[same_text(products_df["Product"], product)].empty:
        messagebox.showinfo("Already Added", "This item name is already added in product/service master")
        return
    if not products_df[same_text(products_df["HSN"], hsn)].empty:
        messagebox.showinfo("Already Added", "This HSN/SAC code is already added in product/service master")
        return

    row = {
        "Product": product,
        "Item Type": item_type,
        "HSN": hsn,
        "Unit": unit,
        "Rate": rate,
    }
    products_df = pd.concat([products_df, pd.DataFrame([row])], ignore_index=True)
    messagebox.showinfo("Added", "Product/service added to master")

    save_master()
    refresh_product_values()


def set_readonly(entry, value):
    entry.config(state="normal")
    entry.delete(0, tk.END)
    entry.insert(0, value)
    entry.config(state="readonly")


def compute_item_tax_values(taxable, rate, tax_type):
    if tax_type == "Intra-State":
        cgst = round(taxable * rate / 200, 2)
        sgst = round(taxable * rate / 200, 2)
        igst = 0.0
    else:
        igst = round(taxable * rate / 100, 2)
        cgst = 0.0
        sgst = 0.0
    total = round(taxable + igst + cgst + sgst, 2)
    return igst, cgst, sgst, total


def update_supply_type_controls(event=None):
    party_gstn = normalize_gstn(party_gstn_var.get())
    client_code = normalize_gstn(client_gstn_var.get())[:2]

    if party_gstn == "URP":
        supply_type_combo.config(values=SUPPLY_TYPES_URP, state="readonly")
        if supply_type_var.get() not in SUPPLY_TYPES_URP:
            supply_type_var.set("B2C-1")
        if supply_type_var.get() == "B2C-2":
            party_state_combo.config(state="readonly")
        else:
            party_state_combo.config(state="disabled")
            party_state_var.set(state_option_from_code(client_code))
    else:
        supply_type_combo.config(values=SUPPLY_TYPES_REGISTERED, state="readonly")
        supply_type_var.set("B2B")
        state_code = party_gstn[:2] if len(party_gstn) >= 2 and party_gstn[:2].isdigit() else ""
        party_state_var.set(state_option_from_code(state_code))
        party_state_combo.config(state="disabled")


def notify_place_of_supply_locked(event=None):
    if str(party_state_combo.cget("state")) != "disabled":
        return

    party_gstn = normalize_gstn(party_gstn_var.get())
    if party_gstn == "URP" and supply_type_var.get() == "B2C-1":
        message = "Place of Supply is locked for B2C-1 because it is an intra-state URP supply."
    elif party_gstn == "URP":
        message = "Place of Supply is editable only when Supply Type is B2C-2."
    else:
        message = "Place of Supply is auto-filled from the registered party GSTN."

    status_var.set(message)
    messagebox.showinfo("Place of Supply Locked", message)


def calculate_tax(event=None):
    try:
        client_gstn = validate_gstn(client_gstn_var.get(), "Client GSTN")
        party_gstn = validate_party_gstn_or_urp(party_gstn_var.get())
    except ValueError:
        for entry in [igst_entry, cgst_entry, sgst_entry, total_entry]:
            set_readonly(entry, "")
        return

    update_supply_type_controls()
    supply_type = normalize_text(supply_type_var.get())

    if party_gstn == "URP":
        is_intra_state = supply_type == "B2C-1"
    else:
        is_intra_state = client_gstn[:2] == party_gstn[:2]

    tax_type = "Intra-State" if is_intra_state else "Inter-State"
    tax_type_var.set(tax_type)

    taxable_total = round(
        sum(float(item.get("Taxable Value", 0) or 0) for item in current_items), 2
    )
    igst_total = round(sum(float(item.get("IGST", 0) or 0) for item in current_items), 2)
    cgst_total = round(sum(float(item.get("CGST", 0) or 0) for item in current_items), 2)
    sgst_total = round(sum(float(item.get("SGST", 0) or 0) for item in current_items), 2)
    invoice_total = round(
        sum(float(item.get("Total Value", 0) or 0) for item in current_items), 2
    )

    taxable_var.set(f"{taxable_total:.2f}" if taxable_total else "")
    set_readonly(igst_entry, f"{igst_total:.2f}" if current_items else "")
    set_readonly(cgst_entry, f"{cgst_total:.2f}" if current_items else "")
    set_readonly(sgst_entry, f"{sgst_total:.2f}" if current_items else "")
    set_readonly(total_entry, f"{invoice_total:.2f}" if current_items else "")


def show_text_dialog(title, body):
    dialog = tk.Toplevel(root)
    dialog.title(title)
    dialog.configure(bg="#f2f2f2")
    dialog.transient(root)
    dialog.geometry("760x560")
    dialog.minsize(620, 420)

    outer = tk.Frame(dialog, bg="#f2f2f2", padx=12, pady=12)
    outer.pack(fill="both", expand=True)
    outer.rowconfigure(1, weight=1)
    outer.columnconfigure(0, weight=1)

    tk.Label(
        outer,
        text=title,
        font=("Arial", 14, "bold"),
        bg="#f2f2f2",
        anchor="w",
    ).grid(row=0, column=0, sticky="ew", pady=(0, 8))

    text_widget = scrolledtext.ScrolledText(
        outer,
        wrap="word",
        font=("Arial", 10),
        padx=10,
        pady=10,
    )
    text_widget.grid(row=1, column=0, sticky="nsew")
    text_widget.insert("1.0", body)
    text_widget.config(state="disabled")

    tk.Button(
        outer,
        text="Close",
        command=dialog.destroy,
        bg="#3f5f8f",
        fg="white",
        width=12,
    ).grid(row=2, column=0, sticky="e", pady=(10, 0))

    dialog.focus_set()


def show_introduction():
    show_text_dialog("INTRODUCTION", INTRODUCTION_TEXT)


def show_about():
    show_text_dialog("About Sales Soldier", INTRODUCTION_TEXT)


def show_help():
    show_text_dialog("Help", HELP_TEXT)


def update_quantity_controls(event=None):
    item_type = normalize_text(item_type_var.get()) or "Goods"
    if item_type == "Service":
        quantity_var.set("")
        unit_var.set("")
        quantity_entry.config(state="disabled")
        unit_combo.config(state="disabled")
        status_var.set("Service selected: quantity and unit are not required")
    else:
        quantity_entry.config(state="normal")
        unit_combo.config(state="readonly")
        if not quantity_var.get():
            quantity_var.set("1")


def clear_item_entry(keep_master=False):
    global editing_item_index
    if not keep_master:
        product_var.set("")
        hsn_var.set("")
        rate_var.set("")
        unit_var.set("")
        item_type_var.set("Goods")
    quantity_var.set("1")
    editing_item_index = None
    update_quantity_controls()


def collect_current_item():
    product = normalize_text(product_var.get())
    item_type = normalize_text(item_type_var.get()) or "Goods"
    hsn = normalize_text(hsn_var.get())
    unit = normalize_text(unit_var.get()).upper()
    taxable = to_float(item_taxable_var.get(), "Item taxable value")
    rate = to_float(rate_var.get(), "GST rate")
    tax_type = normalize_text(tax_type_var.get())

    if not product:
        raise ValueError("Enter product/service")
    if item_type not in ITEM_TYPES:
        raise ValueError("Select Goods or Service")
    if not hsn:
        raise ValueError("Enter HSN / SAC")
    if item_type == "Goods":
        quantity = to_float(quantity_var.get(), "Quantity")
        if quantity <= 0:
            raise ValueError("Quantity must be greater than zero")
        if not unit:
            raise ValueError("Select quantity unit for goods")
    else:
        quantity = ""
        unit = ""
    if taxable < 0:
        raise ValueError("Item taxable value cannot be negative")
    if rate < 0:
        raise ValueError("GST rate cannot be negative")
    if not tax_type:
        raise ValueError("Enter invoice details before adding item")

    igst, cgst, sgst, total_value = compute_item_tax_values(taxable, rate, tax_type)
    return {
        "Product/Service": product,
        "Item Type": item_type,
        "HSN": hsn,
        "Quantity": quantity,
        "Unit": unit,
        "Taxable Value": round(taxable, 2),
        "GST Rate": rate,
        "IGST": igst,
        "CGST": cgst,
        "SGST": sgst,
        "Total Value": total_value,
    }


def refresh_item_list():
    item_tree.delete(*item_tree.get_children())
    for index, item in enumerate(current_items):
        item_tree.insert(
            "",
            "end",
            iid=str(index),
            values=(item.get("Product/Service", ""), item.get("Taxable Value", "")),
        )
    calculate_tax()


def add_or_update_item():
    global editing_item_index
    try:
        item = collect_current_item()
    except ValueError as exc:
        messagebox.showerror("Item Error", str(exc))
        return

    if editing_item_index is None:
        current_items.append(item)
        status_var.set("Item added to invoice")
    else:
        current_items[editing_item_index] = item
        status_var.set("Item updated in invoice")

    editing_item_index = None
    refresh_item_list()
    item_taxable_var.set("")
    clear_item_entry(keep_master=False)


def on_item_select(event=None):
    global editing_item_index
    selected = item_tree.selection()
    if not selected:
        return
    editing_item_index = int(selected[0])
    item = current_items[editing_item_index]
    product_var.set(item.get("Product/Service", ""))
    item_type_var.set(item.get("Item Type", "Goods"))
    hsn_var.set(item.get("HSN", ""))
    quantity_var.set(item.get("Quantity", "1"))
    unit_var.set(item.get("Unit", ""))
    rate_var.set(item.get("GST Rate", ""))
    item_taxable_var.set(item.get("Taxable Value", ""))
    update_quantity_controls()


def delete_selected_item():
    global editing_item_index
    selected = item_tree.selection()
    if not selected:
        return
    index = int(selected[0])
    del current_items[index]
    editing_item_index = None
    refresh_item_list()
    item_taxable_var.set("")
    clear_item_entry(keep_master=False)


def format_invoice_date_field(event=None):
    if not normalize_text(invoice_date_var.get()):
        return
    try:
        parsed = parse_invoice_date(invoice_date_var.get())
        validate_invoice_date_in_month(parsed, month_var.get())
        invoice_date_var.set(parsed)
    except ValueError as exc:
        messagebox.showerror("Invalid Invoice Date", str(exc))
        invoice_date_var.set("")
        invoice_date_entry.focus_set()


def update_month_options(event=None):
    options = get_month_options(fy_var.get())
    month_combo["values"] = options
    if month_var.get() not in options:
        month_var.set("")


def invoice_no_exists(invoice_no):
    invoice_no = normalize_text(invoice_no)
    return any(
        normalize_text(invoice.get("Invoice No")) == invoice_no and index != editing_index
        for index, invoice in enumerate(invoices)
    )


def set_invoice_detail_state(enabled):
    state = "normal" if enabled else "disabled"
    readonly_state = "readonly" if enabled else "disabled"
    for widget in invoice_detail_widgets:
        if widget in [supply_type_combo, item_type_combo, unit_combo]:
            widget.config(state=readonly_state)
        elif widget == party_state_combo:
            update_supply_type_controls()
            if not enabled:
                widget.config(state="disabled")
        else:
            widget.config(state=state)

    for widget in invoice_detail_action_widgets:
        widget.config(state=state)

    if enabled:
        update_supply_type_controls()
        update_quantity_controls()


def set_cancelled_mode(cancelled):
    global current_items, editing_item_index
    invoice_status_var.set("Cancelled" if cancelled else "Active")
    if cancelled:
        current_items = []
        editing_item_index = None
        for var in [
            invoice_date_var,
            party_gstn_var,
            party_name_var,
            party_state_var,
            supply_type_var,
            tax_type_var,
            product_var,
            item_type_var,
            hsn_var,
            quantity_var,
            unit_var,
            item_taxable_var,
            taxable_var,
            rate_var,
        ]:
            var.set("")
        party_name_var.set("Cancelled")
        set_readonly(igst_entry, "")
        set_readonly(cgst_entry, "")
        set_readonly(sgst_entry, "")
        set_readonly(total_entry, "")
        refresh_item_list()
        set_invoice_detail_state(False)
        status_var.set("Invoice marked as cancelled")
    else:
        set_invoice_detail_state(True)
        item_type_var.set(item_type_var.get() or "Goods")
        update_supply_type_controls()
        update_quantity_controls()


def cancel_current_invoice():
    invoice_no = normalize_text(invoice_no_var.get())
    if not invoice_no:
        messagebox.showerror("Cancel Invoice", "Enter invoice number first, then click Cancel")
        invoice_no_entry.focus_set()
        return

    if invoice_no_exists(invoice_no):
        messagebox.showerror(
            "Duplicate Invoice No",
            "Invoice number already exists. Change the invoice number before cancelling.",
        )
        invoice_no_entry.focus_set()
        return

    set_cancelled_mode(True)


def validate_invoice_number_live(event=None):
    global duplicate_invoice_warned
    invoice_no = normalize_text(invoice_no_var.get())
    duplicate = bool(invoice_no) and invoice_no_exists(invoice_no)

    if duplicate:
        set_invoice_detail_state(False)
        message = "Invoice number already exists. Change the invoice number to continue."
        status_var.set(message)
        if not duplicate_invoice_warned:
            duplicate_invoice_warned = True
            messagebox.showerror("Duplicate Invoice No", message)
    else:
        duplicate_invoice_warned = False
        if invoice_status_var.get() != "Cancelled":
            set_invoice_detail_state(True)
        if invoice_no:
            status_var.set("Invoice number accepted")


def collect_invoice():
    invoice_status = normalize_text(invoice_status_var.get()) or "Active"
    client_name = normalize_text(client_name_var.get())
    client_gstn = normalize_gstn(client_gstn_var.get())
    financial_year = normalize_text(fy_var.get())
    month = normalize_text(month_var.get())
    invoice_no = normalize_text(invoice_no_var.get())

    if not financial_year:
        raise ValueError("Select financial year")
    if not month:
        raise ValueError("Select month")
    if not invoice_no:
        raise ValueError("Enter invoice number")
    if not client_name:
        raise ValueError("Enter client name")
    if not client_gstn:
        raise ValueError("Enter client GSTN")
    if invoice_status == "Cancelled":
        duplicate = [
            i
            for i, invoice in enumerate(invoices)
            if invoice.get("Invoice No") == invoice_no and i != editing_index
        ]
        if duplicate:
            raise ValueError("Invoice number already exists in this month")

        return {
            "Financial Year": financial_year,
            "Month": month,
            "Invoice Status": "Cancelled",
            "Client Name": "Cancelled",
            "Client GSTN": "",
            "Invoice No": invoice_no,
            "Invoice Date": "",
            "Party GSTN": "",
            "Party Name": "Cancelled",
            "Party State Code": "",
            "Place of Supply": "",
            "Supply Type": "",
            "Tax Type": "",
            "Product/Service": "",
            "Item Type": "",
            "HSN": "",
            "Quantity": "",
            "Unit": "",
            "Taxable Value": "",
            "GST Rate": "",
            "IGST": "",
            "CGST": "",
            "SGST": "",
            "Total Invoice Value": "",
            "Items": [],
        }

    client_gstn = validate_gstn(client_gstn, "Client GSTN")
    invoice_date = parse_invoice_date(invoice_date_var.get())
    validate_invoice_date_in_month(invoice_date, month)
    party_gstn = validate_party_gstn_or_urp(party_gstn_var.get())
    party_name = normalize_text(party_name_var.get())
    party_state_code = state_code_from_option(party_state_var.get())
    place_of_supply = normalize_text(party_state_var.get())
    supply_type = normalize_text(supply_type_var.get())
    tax_type = normalize_text(tax_type_var.get())

    if not invoice_date:
        raise ValueError("Enter invoice date")
    if not party_name:
        raise ValueError("Enter party name")
    if supply_type not in ["B2B", "B2C-1", "B2C-2"]:
        raise ValueError("Select valid supply type")
    if party_gstn == "URP" and supply_type == "B2B":
        raise ValueError("URP invoice must be B2C-1 or B2C-2")
    if party_gstn != "URP" and supply_type != "B2B":
        raise ValueError("Registered GSTN invoice must be B2B")
    if not party_state_code:
        raise ValueError("Select place of supply/state")
    if not current_items:
        raise ValueError("Add at least one item")

    calculate_tax()

    taxable = sum(float(item.get("Taxable Value", 0) or 0) for item in current_items)
    igst = sum(float(item.get("IGST", 0) or 0) for item in current_items)
    cgst = sum(float(item.get("CGST", 0) or 0) for item in current_items)
    sgst = sum(float(item.get("SGST", 0) or 0) for item in current_items)
    total = sum(float(item.get("Total Value", 0) or 0) for item in current_items)
    tax_type = normalize_text(tax_type_var.get())

    duplicate = [
        i
        for i, invoice in enumerate(invoices)
        if invoice.get("Invoice No") == invoice_no and i != editing_index
    ]
    if duplicate:
        raise ValueError("Invoice number already exists in this month")

    return {
        "Financial Year": financial_year,
        "Month": month,
        "Invoice Status": "Active",
        "Client Name": client_name,
        "Client GSTN": client_gstn,
        "Invoice No": invoice_no,
        "Invoice Date": invoice_date,
        "Party GSTN": party_gstn,
        "Party Name": party_name,
        "Party State Code": party_state_code,
        "Place of Supply": place_of_supply,
        "Supply Type": supply_type,
        "Tax Type": tax_type,
        "Product/Service": "",
        "Item Type": "",
        "HSN": "",
        "Quantity": "",
        "Unit": "",
        "Taxable Value": round(taxable, 2),
        "GST Rate": "",
        "IGST": round(igst, 2),
        "CGST": round(cgst, 2),
        "SGST": round(sgst, 2),
        "Total Invoice Value": round(total, 2),
        "Items": [dict(item) for item in current_items],
    }


def save_invoice():
    global editing_index, current_items
    try:
        invoice = collect_invoice()
    except ValueError as exc:
        messagebox.showerror("Check Entry", str(exc))
        return

    if editing_index is None:
        invoices.append(invoice)
        status_var.set("Invoice added")
    else:
        invoices[editing_index] = invoice
        status_var.set("Invoice updated")

    refresh_tree()
    save_draft()
    current_items = []
    clear_form(keep_header=True)


def load_invoice_to_form(index):
    global current_items
    invoice = invoices[index]
    fy_var.set(invoice.get("Financial Year", ""))
    update_month_options()
    month_var.set(invoice.get("Month", ""))
    invoice_status_var.set(invoice.get("Invoice Status", "Active") or "Active")
    invoice_no_var.set(invoice.get("Invoice No", ""))
    invoice_date_var.set(invoice.get("Invoice Date", ""))
    party_gstn_var.set(invoice.get("Party GSTN", ""))
    party_name_var.set(invoice.get("Party Name", ""))
    party_state_var.set(invoice.get("Place of Supply", ""))
    supply_type_var.set(invoice.get("Supply Type", "B2B") or "B2B")
    tax_type_var.set(invoice.get("Tax Type", ""))
    taxable_var.set(invoice.get("Taxable Value", ""))
    current_items = [dict(item) for item in invoice.get("Items", [])]
    if not current_items and normalize_text(invoice.get("Product/Service", "")):
        current_items = [
            {
                "Product/Service": invoice.get("Product/Service", ""),
                "Item Type": invoice.get("Item Type", ""),
                "HSN": invoice.get("HSN", ""),
                "Quantity": invoice.get("Quantity", ""),
                "Unit": invoice.get("Unit", ""),
                "Taxable Value": invoice.get("Taxable Value", ""),
                "GST Rate": invoice.get("GST Rate", ""),
                "IGST": invoice.get("IGST", ""),
                "CGST": invoice.get("CGST", ""),
                "SGST": invoice.get("SGST", ""),
                "Total Value": invoice.get("Total Invoice Value", ""),
            }
        ]
    refresh_item_list()
    item_taxable_var.set("")
    clear_item_entry()
    if invoice_status_var.get() == "Cancelled":
        set_cancelled_mode(True)
    else:
        update_quantity_controls()
        update_supply_type_controls()

    set_readonly(igst_entry, invoice.get("IGST", ""))
    set_readonly(cgst_entry, invoice.get("CGST", ""))
    set_readonly(sgst_entry, invoice.get("SGST", ""))
    set_readonly(total_entry, invoice.get("Total Invoice Value", ""))


def on_select(index=None):
    global editing_index
    if index is None:
        return
    editing_index = int(index)
    load_invoice_to_form(editing_index)
    refresh_tree()
    update_button_state()
    status_var.set("Editing selected invoice")


def delete_invoice():
    global editing_index
    if editing_index is None:
        return

    invoice_no = invoices[editing_index].get("Invoice No", "")
    confirm = messagebox.askyesno("Delete Invoice", f"Delete invoice {invoice_no}?")
    if not confirm:
        return

    del invoices[editing_index]
    editing_index = None
    refresh_tree()
    save_draft()
    clear_form(keep_header=True)
    status_var.set("Invoice deleted")


def clear_form(keep_header=False):
    global editing_index, current_items, editing_item_index
    editing_index = None
    editing_item_index = None
    current_items = []

    header_values = {"fy": fy_var.get(), "month": month_var.get()}

    for var in [
        fy_var,
        month_var,
        invoice_no_var,
        invoice_date_var,
        party_gstn_var,
        party_name_var,
        party_state_var,
        supply_type_var,
        tax_type_var,
        product_var,
        item_type_var,
        hsn_var,
        quantity_var,
        unit_var,
        item_taxable_var,
        taxable_var,
        rate_var,
    ]:
        var.set("")

    invoice_status_var.set("Active")
    item_type_var.set("Goods")
    quantity_var.set("1")
    supply_type_var.set("B2B")

    if keep_header:
        fy_var.set(header_values["fy"])
        month_var.set(header_values["month"])

    for entry in [igst_entry, cgst_entry, sgst_entry, total_entry]:
        set_readonly(entry, "")

    refresh_item_list()
    set_invoice_detail_state(True)
    update_quantity_controls()
    update_button_state()


def write_export_workbook(file, sheets):
    workbook = Workbook()
    default_sheet = workbook.active

    for index, (sheet_name, dataframe) in enumerate(sheets):
        if index == 0:
            worksheet = default_sheet
            worksheet.title = sheet_name
        else:
            worksheet = workbook.create_sheet(title=sheet_name)

        for row in dataframe_to_rows(dataframe, index=False, header=True):
            worksheet.append(["" if pd.isna(value) else value for value in row])

        for column_cells in worksheet.columns:
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(max_length + 2, 12),
                35,
            )

    workbook.save(file)


def get_invoice_summary_rows():
    rows = []
    for invoice in invoices:
        row = {column: invoice.get(column, "") for column in INVOICE_COLUMNS}
        row.pop("Items", None)
        rows.append(row)
    return rows


def get_flat_invoice_rows():
    rows = []
    for invoice in invoices:
        base = {column: invoice.get(column, "") for column in INVOICE_COLUMNS}
        items = invoice.get("Items", [])
        if invoice.get("Invoice Status") == "Cancelled" or not items:
            base["Items"] = []
            rows.append(base)
            continue
        for item in items:
            row = dict(base)
            row["Product/Service"] = item.get("Product/Service", "")
            row["Item Type"] = item.get("Item Type", "")
            row["HSN"] = item.get("HSN", "")
            row["Quantity"] = item.get("Quantity", "")
            row["Unit"] = item.get("Unit", "")
            row["GST Rate"] = item.get("GST Rate", "")
            row["Taxable Value"] = item.get("Taxable Value", "")
            row["IGST"] = item.get("IGST", "")
            row["CGST"] = item.get("CGST", "")
            row["SGST"] = item.get("SGST", "")
            row["Total Invoice Value"] = item.get("Total Value", "")
            rows.append(row)
    return rows


def format_place_of_supply(value):
    value = normalize_text(value)
    if " - " in value:
        code, state = value.split(" - ", 1)
        return f"{code}-{state}"
    return value


def format_uqc(value):
    value = normalize_text(value).upper()
    return UQC_DISPLAY.get(value, value)


def build_b2b_output(summary_df, item_df):
    summary_b2b_df = summary_df[
        (summary_df["Invoice Status"] != "Cancelled") & (summary_df["Supply Type"] == "B2B")
    ].copy()
    item_b2b_df = item_df[
        (item_df["Invoice Status"] != "Cancelled") & (item_df["Supply Type"] == "B2B")
    ].copy()

    output_columns = [
        "GSTIN",
        "Receiver Name",
        "Invoice Number",
        "Invoice date",
        "Invoice Value",
        "Place Of Supply",
        "Reverse Charge",
        "Applicable % of Tax Rate",
        "Invoice Type",
        "E-Commerce GSTIN",
        "Rate",
        "Taxable Value",
        "Cess Amount",
    ]

    if summary_b2b_df.empty or item_b2b_df.empty:
        return pd.DataFrame(columns=output_columns)

    item_b2b_df["Rate"] = pd.to_numeric(item_b2b_df["GST Rate"], errors="coerce")
    item_b2b_df["Taxable Value"] = pd.to_numeric(
        item_b2b_df["Taxable Value"], errors="coerce"
    ).fillna(0)

    grouped_items = (
        item_b2b_df.groupby(["Invoice No", "Rate"], as_index=False, dropna=False)
        .agg({"Taxable Value": "sum"})
    )
    grouped_items["_invoice_sort"] = grouped_items["Invoice No"].map(natural_sort_tuple)
    grouped_items = grouped_items.sort_values(["_invoice_sort", "Rate"]).drop(
        columns=["_invoice_sort"]
    )

    summary_lookup = summary_b2b_df[
        [
            "Invoice No",
            "Party GSTN",
            "Party Name",
            "Invoice Date",
            "Total Invoice Value",
            "Place of Supply",
        ]
    ].drop_duplicates(subset=["Invoice No"])

    merged_df = grouped_items.merge(summary_lookup, on="Invoice No", how="left")

    output_df = pd.DataFrame(
        {
            "GSTIN": merged_df["Party GSTN"],
            "Receiver Name": merged_df["Party Name"],
            "Invoice Number": merged_df["Invoice No"],
            "Invoice date": merged_df["Invoice Date"],
            "Invoice Value": pd.to_numeric(
                merged_df["Total Invoice Value"], errors="coerce"
            ),
            "Place Of Supply": merged_df["Place of Supply"].apply(format_place_of_supply),
            "Reverse Charge": "N",
            "Applicable % of Tax Rate": 0,
            "Invoice Type": "Regular B2B",
            "E-Commerce GSTIN": "",
            "Rate": merged_df["Rate"],
            "Taxable Value": merged_df["Taxable Value"],
            "Cess Amount": "",
        }
    )
    return output_df[output_columns]


def build_b2c_output(invoice_df):
    b2c_df = invoice_df[
        (invoice_df["Invoice Status"] != "Cancelled")
        & (invoice_df["Supply Type"].isin(["B2C-1", "B2C-2"]))
    ].copy()

    output_columns = [
        "Type",
        "Place Of Supply",
        "Applicable % of Tax Rate",
        "Rate",
        "Taxable Value",
        "Cess Amount",
        "E-Commerce GSTIN",
    ]

    if b2c_df.empty:
        return pd.DataFrame(columns=output_columns)

    b2c_df["Place Of Supply"] = b2c_df["Place of Supply"].apply(format_place_of_supply)
    b2c_df["Rate"] = pd.to_numeric(b2c_df["GST Rate"], errors="coerce")
    b2c_df["Taxable Value"] = pd.to_numeric(
        b2c_df["Taxable Value"], errors="coerce"
    ).fillna(0)

    grouped_df = (
        b2c_df.groupby(["Place Of Supply", "Rate"], as_index=False, dropna=False)
        .agg({"Taxable Value": "sum"})
        .sort_values(["Place Of Supply", "Rate"])
    )

    output_df = pd.DataFrame(
        {
            "Type": "OE",
            "Place Of Supply": grouped_df["Place Of Supply"],
            "Applicable % of Tax Rate": 0,
            "Rate": grouped_df["Rate"],
            "Taxable Value": grouped_df["Taxable Value"],
            "Cess Amount": 0,
            "E-Commerce GSTIN": "",
        }
    )
    return output_df[output_columns]


def build_hsn_b2b_output(invoice_df):
    b2b_df = invoice_df[
        (invoice_df["Invoice Status"] != "Cancelled") & (invoice_df["Supply Type"] == "B2B")
    ].copy()

    output_columns = [
        "HSN",
        "Description",
        "UQC",
        "Total Quantity",
        "Total Value",
        "Rate",
        "Taxable Value",
        "Integrated Tax Amount",
        "Central Tax Amount",
        "State/UT Tax Amount",
        "Cess Amount",
    ]

    if b2b_df.empty:
        return pd.DataFrame(columns=output_columns)

    b2b_df["Quantity"] = pd.to_numeric(b2b_df["Quantity"], errors="coerce").fillna(0)
    b2b_df["Rate"] = pd.to_numeric(b2b_df["GST Rate"], errors="coerce")
    b2b_df["Taxable Value"] = pd.to_numeric(
        b2b_df["Taxable Value"], errors="coerce"
    ).fillna(0)
    b2b_df["IGST"] = pd.to_numeric(b2b_df["IGST"], errors="coerce").fillna(0)
    b2b_df["CGST"] = pd.to_numeric(b2b_df["CGST"], errors="coerce").fillna(0)
    b2b_df["SGST"] = pd.to_numeric(b2b_df["SGST"], errors="coerce").fillna(0)
    b2b_df["UQC"] = b2b_df["Unit"].apply(format_uqc)

    grouped_df = (
        b2b_df.groupby(["HSN", "UQC", "Rate"], as_index=False, dropna=False)
        .agg(
            {
                "Quantity": "sum",
                "Taxable Value": "sum",
                "IGST": "sum",
                "CGST": "sum",
                "SGST": "sum",
            }
        )
        .sort_values(["HSN", "Rate", "UQC"])
    )

    output_df = pd.DataFrame(
        {
            "HSN": grouped_df["HSN"],
            "Description": "",
            "UQC": grouped_df["UQC"],
            "Total Quantity": grouped_df["Quantity"],
            "Total Value": 0,
            "Rate": grouped_df["Rate"],
            "Taxable Value": grouped_df["Taxable Value"],
            "Integrated Tax Amount": grouped_df["IGST"],
            "Central Tax Amount": grouped_df["CGST"],
            "State/UT Tax Amount": grouped_df["SGST"],
            "Cess Amount": 0,
        }
    )
    return output_df[output_columns]


def build_hsn_b2c_output(invoice_df):
    b2c_df = invoice_df[
        (invoice_df["Invoice Status"] != "Cancelled")
        & (invoice_df["Supply Type"].isin(["B2C-1", "B2C-2"]))
    ].copy()

    output_columns = [
        "HSN",
        "Description",
        "UQC",
        "Total Quantity",
        "Total Value",
        "Rate",
        "Taxable Value",
        "Integrated Tax Amount",
        "Central Tax Amount",
        "State/UT Tax Amount",
        "Cess Amount",
    ]

    if b2c_df.empty:
        return pd.DataFrame(columns=output_columns)

    b2c_df["Quantity"] = pd.to_numeric(b2c_df["Quantity"], errors="coerce").fillna(0)
    b2c_df["Rate"] = pd.to_numeric(b2c_df["GST Rate"], errors="coerce")
    b2c_df["Taxable Value"] = pd.to_numeric(
        b2c_df["Taxable Value"], errors="coerce"
    ).fillna(0)
    b2c_df["IGST"] = pd.to_numeric(b2c_df["IGST"], errors="coerce").fillna(0)
    b2c_df["CGST"] = pd.to_numeric(b2c_df["CGST"], errors="coerce").fillna(0)
    b2c_df["SGST"] = pd.to_numeric(b2c_df["SGST"], errors="coerce").fillna(0)
    b2c_df["UQC"] = b2c_df["Unit"].apply(format_uqc)
    b2c_df["Place Of Supply"] = b2c_df["Place of Supply"].apply(format_place_of_supply)

    grouped_df = (
        b2c_df.groupby(
            ["Place Of Supply", "HSN", "UQC", "Rate"], as_index=False, dropna=False
        )
        .agg(
            {
                "Quantity": "sum",
                "Taxable Value": "sum",
                "IGST": "sum",
                "CGST": "sum",
                "SGST": "sum",
            }
        )
        .sort_values(["Place Of Supply", "HSN", "Rate", "UQC"])
    )

    output_df = pd.DataFrame(
        {
            "HSN": grouped_df["HSN"],
            "Description": "",
            "UQC": grouped_df["UQC"],
            "Total Quantity": grouped_df["Quantity"],
            "Total Value": 0,
            "Rate": grouped_df["Rate"],
            "Taxable Value": grouped_df["Taxable Value"],
            "Integrated Tax Amount": grouped_df["IGST"],
            "Central Tax Amount": grouped_df["CGST"],
            "State/UT Tax Amount": grouped_df["SGST"],
            "Cess Amount": 0,
        }
    )
    return output_df[output_columns]


def build_output_document(invoice_df):
    output_columns = [
        "Nature of Document",
        "Sr. No. From",
        "Sr. No. To",
        "Total Number",
        "Cancelled",
    ]

    if invoice_df.empty:
        return pd.DataFrame(columns=output_columns)

    invoice_numbers = [
        normalize_text(value)
        for value in invoice_df["Invoice No"].tolist()
        if normalize_text(value)
    ]
    sorted_invoice_numbers = sorted(invoice_numbers, key=natural_sort_key)

    cancelled_count = int(
        (
            invoice_df["Invoice Status"].fillna("").astype(str).str.strip().str.lower()
            == "cancelled"
        ).sum()
    )

    return pd.DataFrame(
        [
            {
                "Nature of Document": "Invoices for outward supply",
                "Sr. No. From": sorted_invoice_numbers[0] if sorted_invoice_numbers else "",
                "Sr. No. To": sorted_invoice_numbers[-1] if sorted_invoice_numbers else "",
                "Total Number": len(sorted_invoice_numbers),
                "Cancelled": cancelled_count,
            }
        ],
        columns=output_columns,
    )


def copy_range_values(source_sheet, start_row, start_col, end_row, end_col, target_sheet, target_row, target_col):
    for row_offset, source_row in enumerate(range(start_row, end_row + 1)):
        for col_offset, source_col in enumerate(range(start_col, end_col + 1)):
            target_sheet.cell(
                row=target_row + row_offset,
                column=target_col + col_offset,
                value=source_sheet.cell(row=source_row, column=source_col).value,
            )


def find_last_data_row(sheet, start_row, start_col, end_col):
    last_data_row = start_row - 1
    for row in range(start_row, sheet.max_row + 1):
        if any(
            sheet.cell(row=row, column=column).value not in [None, ""]
            for column in range(start_col, end_col + 1)
        ):
            last_data_row = row
    return last_data_row


def extract_transfer_rows(source_sheet, end_col):
    last_data_row = find_last_data_row(source_sheet, start_row=2, start_col=1, end_col=end_col)
    rows = []
    if last_data_row < 2:
        return rows

    for row_index in range(2, last_data_row + 1):
        row_values = []
        for column_index in range(1, end_col + 1):
            value = source_sheet.cell(row=row_index, column=column_index).value
            if pd.isna(value):
                value = None
            row_values.append(value)
        rows.append(row_values)
    return rows


def transfer_with_excel_com(target_template_file, transfer_payload):
    payload_handle, payload_path = tempfile.mkstemp(
        suffix=".json", prefix="sales_soldier_transfer_", dir=APP_DIR
    )
    script_handle, script_path = tempfile.mkstemp(
        suffix=".ps1", prefix="sales_soldier_transfer_", dir=APP_DIR
    )

    os.close(payload_handle)
    os.close(script_handle)

    script_content = r"""
param(
    [string]$WorkbookPath,
    [string]$PayloadPath
)

$ErrorActionPreference = 'Stop'
$excel = $null
$workbook = $null

try {
    $payload = Get-Content -LiteralPath $PayloadPath -Raw | ConvertFrom-Json
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $workbook = $excel.Workbooks.Open($WorkbookPath)

    foreach ($mapping in $payload.mappings) {
        $sheet = $workbook.Worksheets.Item($mapping.target_sheet)
        $startRow = [int]$mapping.target_row
        $startCol = [int]$mapping.target_col
        $endCol = [int]$mapping.end_col
        $clearEndRow = [Math]::Max($startRow, $startRow + 5000)
        $sheet.Range($sheet.Cells($startRow, $startCol), $sheet.Cells($clearEndRow, $endCol)).ClearContents()

        $rowOffset = 0
        foreach ($row in $mapping.rows) {
            $colOffset = 0
            foreach ($cell in $row) {
                $targetCell = $sheet.Cells($startRow + $rowOffset, $startCol + $colOffset)
                if ($null -eq $cell -or $cell -eq "") {
                    $targetCell.Value2 = $null
                } else {
                    $targetCell.Value2 = $cell
                }
                $colOffset++
            }
            $rowOffset++
        }
    }

    $workbook.Save()
    $workbook.Close($true)
    $excel.Quit()
}
finally {
    if ($workbook -ne $null) {
        [System.Runtime.Interopservices.Marshal]::ReleaseComObject($workbook) | Out-Null
    }
    if ($excel -ne $null) {
        [System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null
    }
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}
"""

    try:
        with open(payload_path, "w", encoding="utf-8") as payload_file:
            json.dump(transfer_payload, payload_file, ensure_ascii=True)
        with open(script_path, "w", encoding="utf-8") as script_file:
            script_file.write(script_content)

        result = subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                script_path,
                target_template_file,
                payload_path,
            ],
            capture_output=True,
            text=True,
            check=False,
        )
        if result.returncode != 0:
            error_text = (result.stderr or result.stdout or "").strip()
            if not error_text:
                error_text = "Excel transfer failed"
            raise RuntimeError(error_text)
    finally:
        for temp_path in [payload_path, script_path]:
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except OSError:
                    pass


def transfer_to_gstr1_template():
    try:
        source_final_file = get_latest_final_gstr1_file()
        template_source_file = get_template_source_file()
        target_template_file = get_transfer_template_filename()
    except (ValueError, FileNotFoundError) as exc:
        messagebox.showerror("Transfer Error", str(exc))
        return

    try:
        shutil.copy2(template_source_file, target_template_file)

        source_workbook = load_workbook(source_final_file, data_only=True)

        if "OUTPUT-B2B" not in source_workbook.sheetnames:
            raise ValueError("One or more required output sheets are not available in final GSTR-1 file")

        transfer_maps = [
            ("OUTPUT-B2B", "b2b,sez,de", 13),
            ("OUTPUT-B2C", "b2cs", 7),
            ("OUTPUT-HSN_B2B", "hsn(b2b)", 11),
            ("OUTPUT-HSN_B2C", "hsn(b2c)", 11),
            ("OUTPUT-DOCUMENT", "docs", 5),
        ]

        transfer_payload = {"mappings": []}
        for source_name, target_name, end_col in transfer_maps:
            if source_name not in source_workbook.sheetnames:
                raise ValueError(f"{source_name} sheet not available in final GSTR-1 file")
            source_sheet = source_workbook[source_name]
            transfer_payload["mappings"].append(
                {
                    "source_sheet": source_name,
                    "target_sheet": target_name,
                    "target_row": 5,
                    "target_col": 1,
                    "end_col": end_col,
                    "rows": extract_transfer_rows(source_sheet, end_col),
                }
            )

        transfer_with_excel_com(target_template_file, transfer_payload)
        messagebox.showinfo(
            "Transfer Complete",
            f"Transferred final GSTR-1 values to:\n{os.path.abspath(target_template_file)}",
        )
        status_var.set(f"Transfer complete: {os.path.abspath(target_template_file)}")
    except Exception as exc:
        messagebox.showerror("Transfer Error", str(exc))


def export_final_gstr1_excel():
    if not invoices:
        messagebox.showerror("No Data", "No invoices to export")
        return

    try:
        file = get_final_gstr1_filename()
    except ValueError as exc:
        messagebox.showerror("Missing Month", str(exc))
        return

    temp_file = ""
    try:
        summary_df = pd.DataFrame(get_invoice_summary_rows(), columns=INVOICE_COLUMNS).fillna("")
        main_df = pd.DataFrame(get_flat_invoice_rows(), columns=INVOICE_COLUMNS).fillna("")
        output_b2b_df = build_b2b_output(summary_df, main_df)
        output_b2c_df = build_b2c_output(summary_df)
        output_hsn_b2b_df = build_hsn_b2b_output(main_df)
        output_hsn_b2c_df = build_hsn_b2c_output(main_df)
        output_document_df = build_output_document(summary_df)

        directory = os.path.dirname(os.path.abspath(file)) or APP_DIR
        handle, temp_file = tempfile.mkstemp(
            suffix=".xlsx",
            prefix="~gstr1_final_",
            dir=directory,
        )
        os.close(handle)
        os.remove(temp_file)

        write_export_workbook(
            temp_file,
            [
                ("MAIN SHEET", main_df),
                ("OUTPUT-B2B", output_b2b_df),
                ("OUTPUT-B2C", output_b2c_df),
                ("OUTPUT-HSN_B2B", output_hsn_b2b_df),
                ("OUTPUT-HSN_B2C", output_hsn_b2c_df),
                ("OUTPUT-DOCUMENT", output_document_df),
            ],
        )

        os.replace(temp_file, file)
        messagebox.showinfo(
            "Exported",
            f"Final GSTR-1 exported successfully.\nSaved at:\n{os.path.abspath(file)}",
        )
        status_var.set(f"Final GSTR-1 exported: {os.path.abspath(file)}")
    except Exception as exc:
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except OSError:
                pass
        messagebox.showerror("Final GSTR-1 Export Error", str(exc))


def export_excel():
    if not invoices:
        messagebox.showerror("No Data", "No invoices to export")
        return

    try:
        file = get_export_filename()
    except ValueError as exc:
        messagebox.showerror("Missing Month", str(exc))
        return

    temp_file = ""
    try:
        invoice_df = pd.DataFrame(get_flat_invoice_rows(), columns=INVOICE_COLUMNS).fillna("")
        invoice_export_df = invoice_df.copy()
        numeric_columns = [
            "Quantity",
            "Taxable Value",
            "GST Rate",
            "IGST",
            "CGST",
            "SGST",
            "Total Invoice Value",
        ]
        for column in numeric_columns:
            invoice_df[column] = pd.to_numeric(invoice_df[column], errors="coerce")
            invoice_export_df[column] = pd.to_numeric(
                invoice_export_df[column], errors="coerce"
            )

        hsn_summary_source = invoice_df[invoice_df["Invoice Status"] != "Cancelled"].copy()

        hsn_summary = (
            hsn_summary_source.groupby(
                ["HSN", "Product/Service", "Item Type", "Unit", "GST Rate"],
                as_index=False,
                dropna=False,
            )
            .agg(
                {
                    "Quantity": "sum",
                    "Taxable Value": "sum",
                    "IGST": "sum",
                    "CGST": "sum",
                    "SGST": "sum",
                    "Total Invoice Value": "sum",
                }
            )
            .rename(columns={"Product/Service": "Description"})
        )
        hsn_summary["Quantity"] = hsn_summary["Quantity"].astype(object)
        hsn_summary["Unit"] = hsn_summary["Unit"].astype(object)
        hsn_summary["Quantity"] = hsn_summary["Quantity"].where(
            hsn_summary["Quantity"].notna(), ""
        )
        hsn_summary.loc[hsn_summary["Item Type"] == "Service", ["Quantity", "Unit"]] = ""

        directory = os.path.dirname(os.path.abspath(file)) or APP_DIR
        handle, temp_file = tempfile.mkstemp(
            suffix=".xlsx",
            prefix="~gstr1_export_",
            dir=directory,
        )
        os.close(handle)
        os.remove(temp_file)

        write_export_workbook(
            temp_file,
            [
                ("Invoices", invoice_export_df),
                ("HSN Summary", hsn_summary),
            ],
        )

        os.replace(temp_file, file)
        messagebox.showinfo(
            "Exported",
            f"Exported successfully.\nSaved at:\n{os.path.abspath(file)}",
        )
        status_var.set(f"Exported: {os.path.abspath(file)}")
    except Exception as exc:
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except OSError:
                pass
        messagebox.showerror("Export Error", str(exc))

def make_label(parent, text, row, column):
    label = tk.Label(parent, text=text, bg="#f2f2f2", anchor="w", font=("Arial", 9))
    label.grid(row=row, column=column, sticky="w", padx=(8, 5), pady=4)
    return label


def make_entry(parent, variable, row, column, width=28):
    entry = tk.Entry(parent, textvariable=variable, width=width, font=("Arial", 9))
    entry.grid(row=row, column=column, sticky="ew", padx=(0, 10), pady=4)
    return entry


def create_app_icon():
    icon = tk.PhotoImage(width=32, height=32)
    icon.put("#0f5132", to=(0, 0, 31, 31))
    icon.put("#d1e7dd", to=(4, 4, 27, 27))
    icon.put("#198754", to=(8, 8, 23, 23))
    icon.put("#ffffff", to=(12, 6, 19, 25))
    icon.put("#198754", to=(14, 8, 17, 23))
    return icon


parties_df, products_df = load_master()

root = tk.Tk()
root.title("Sales Soldier")
root.geometry("1320x820")
root.minsize(1180, 760)
root.configure(bg="#2c4d7a")
app_icon = create_app_icon()
root.iconphoto(True, app_icon)

client_name_var = tk.StringVar()
client_gstn_var = tk.StringVar()
fy_var = tk.StringVar()
month_var = tk.StringVar()
invoice_status_var = tk.StringVar(value="Active")
invoice_no_var = tk.StringVar()
invoice_date_var = tk.StringVar()
party_gstn_var = tk.StringVar()
party_name_var = tk.StringVar()
party_state_var = tk.StringVar()
supply_type_var = tk.StringVar(value="B2B")
tax_type_var = tk.StringVar()
product_var = tk.StringVar()
item_type_var = tk.StringVar(value="Goods")
hsn_var = tk.StringVar()
quantity_var = tk.StringVar(value="1")
unit_var = tk.StringVar()
item_taxable_var = tk.StringVar()
taxable_var = tk.StringVar()
rate_var = tk.StringVar()
igst_var = tk.StringVar()
cgst_var = tk.StringVar()
sgst_var = tk.StringVar()
total_var = tk.StringVar()
status_var = tk.StringVar(value="Ready")

style = ttk.Style()
style.theme_use("default")
style.configure("Treeview", rowheight=22)
style.configure("Treeview.Heading", font=("Arial", 9, "bold"))

main = tk.Frame(root, bg="#f2f2f2")
main.pack(fill="both", expand=True, padx=8, pady=8)
main.columnconfigure(0, weight=1)
main.columnconfigure(1, weight=0)
main.rowconfigure(1, weight=1)
main.rowconfigure(2, weight=0)

header = tk.Frame(main, bg="#3f5f8f", height=48)
header.grid(row=0, column=0, columnspan=2, sticky="ew")
header.grid_propagate(False)
for column in range(8):
    header.columnconfigure(column, weight=1)

tk.Label(header, text="CLIENT", fg="white", bg="#3f5f8f").grid(
    row=0, column=0, sticky="e", padx=(8, 4), pady=10
)
client_name_entry = tk.Entry(header, textvariable=client_name_var, width=24)
client_name_entry.grid(
    row=0, column=1, sticky="ew", padx=(0, 16), pady=10
)
tk.Label(header, text="GSTN", fg="white", bg="#3f5f8f").grid(
    row=0, column=2, sticky="e", padx=(8, 4), pady=10
)
client_gstn_entry = tk.Entry(header, textvariable=client_gstn_var, width=22)
client_gstn_entry.grid(row=0, column=3, sticky="ew", padx=(0, 16), pady=10)
tk.Label(header, text="FY", fg="white", bg="#3f5f8f").grid(
    row=0, column=4, sticky="e", padx=(8, 4), pady=10
)
fy_combo = ttk.Combobox(header, textvariable=fy_var, values=FINANCIAL_YEARS, state="readonly", width=14)
fy_combo.grid(row=0, column=5, sticky="ew", padx=(0, 16), pady=10)
fy_combo.bind("<<ComboboxSelected>>", update_month_options)
tk.Label(header, text="MONTH", fg="white", bg="#3f5f8f").grid(
    row=0, column=6, sticky="e", padx=(8, 4), pady=10
)
month_combo = ttk.Combobox(header, textvariable=month_var, state="readonly", width=18)
month_combo.grid(row=0, column=7, sticky="ew", padx=(0, 10), pady=10)

form = tk.Frame(main, bg="#f2f2f2")
form.grid(row=1, column=0, sticky="nsew", padx=(8, 6), pady=8)
for column in [1, 3]:
    form.columnconfigure(column, weight=1)

section_font = ("Arial", 10, "bold")
tk.Label(form, text="Invoice Details", bg="#f2f2f2", font=section_font).grid(
    row=0, column=0, columnspan=4, sticky="w", padx=8, pady=(0, 4)
)

make_label(form, "Invoice number", 1, 0)
invoice_no_entry = make_entry(form, invoice_no_var, 1, 1)
invoice_no_entry.bind("<KeyRelease>", validate_invoice_number_live)
invoice_no_entry.bind("<FocusOut>", validate_invoice_number_live)
make_label(form, "Invoice date", 1, 2)
invoice_date_entry = make_entry(form, invoice_date_var, 1, 3)
invoice_date_entry.bind("<FocusOut>", format_invoice_date_field)

make_label(form, "Party GSTN", 2, 0)
party_gstn_combo = ttk.Combobox(form, textvariable=party_gstn_var, width=26)
party_gstn_combo.grid(row=2, column=1, sticky="ew", padx=(0, 12), pady=6)
party_gstn_combo.bind("<KeyRelease>", filter_party_gstn_options)
party_gstn_combo.bind("<<Paste>>", defer_filter_party_gstn)
party_gstn_combo.bind("<<ComboboxSelected>>", select_party_from_master_by_gstn)
make_label(form, "Supply Type", 2, 2)
supply_type_combo = ttk.Combobox(
    form,
    textvariable=supply_type_var,
    values=SUPPLY_TYPES_REGISTERED,
    state="readonly",
    width=26,
)
supply_type_combo.grid(row=2, column=3, sticky="ew", padx=(0, 12), pady=6)
supply_type_combo.bind("<<ComboboxSelected>>", lambda event: calculate_tax())

make_label(form, "Party name", 3, 0)
party_name_combo = ttk.Combobox(form, textvariable=party_name_var, width=26)
party_name_combo.grid(row=3, column=1, sticky="ew", padx=(0, 12), pady=6)
party_name_combo.bind("<KeyRelease>", filter_party_name_options)
party_name_combo.bind("<<Paste>>", defer_filter_party_name)
party_name_combo.bind("<<ComboboxSelected>>", select_party_from_master_by_name)
make_label(form, "Place of Supply", 3, 2)
party_state_combo = ttk.Combobox(
    form,
    textvariable=party_state_var,
    values=STATE_OPTIONS,
    state="disabled",
    width=26,
)
party_state_combo.grid(row=3, column=3, sticky="ew", padx=(0, 12), pady=6)
party_state_combo.bind("<Button-1>", notify_place_of_supply_locked)
party_state_combo.bind("<FocusIn>", notify_place_of_supply_locked)

add_party_button = tk.Button(
    form,
    text="+ Add Party",
    bg="black",
    fg="white",
    command=add_or_update_party,
)
add_party_button.grid(row=4, column=1, sticky="w", padx=(0, 10), pady=(2, 8))

cancel_invoice_button = tk.Button(
    form,
    text="Cancel",
    bg="#8c1f1f",
    fg="white",
    command=cancel_current_invoice,
)
cancel_invoice_button.grid(row=4, column=2, sticky="w", padx=(0, 10), pady=(2, 8))

tk.Label(form, text="Product / Service", bg="#f2f2f2", font=section_font).grid(
    row=5, column=0, columnspan=4, sticky="w", padx=8, pady=(2, 4)
)

make_label(form, "Product/Service", 6, 0)
product_combo = ttk.Combobox(form, textvariable=product_var, width=26)
product_combo.grid(row=6, column=1, sticky="ew", padx=(0, 12), pady=6)
product_combo.bind("<KeyRelease>", filter_product_options)
product_combo.bind("<<Paste>>", defer_filter_product)
product_combo.bind("<<ComboboxSelected>>", select_product_from_master)

make_label(form, "Goods / Service", 6, 2)
item_type_combo = ttk.Combobox(
    form,
    textvariable=item_type_var,
    values=ITEM_TYPES,
    state="readonly",
    width=26,
)
item_type_combo.grid(row=6, column=3, sticky="ew", padx=(0, 12), pady=6)
item_type_combo.bind("<<ComboboxSelected>>", update_quantity_controls)

make_label(form, "HSN / SAC", 7, 0)
hsn_entry = make_entry(form, hsn_var, 7, 1)
hsn_entry.bind("<KeyRelease>", on_hsn_keyrelease)
make_label(form, "Quantity", 7, 2)
quantity_entry = make_entry(form, quantity_var, 7, 3)

make_label(form, "Unit / UQC", 8, 0)
unit_combo = ttk.Combobox(
    form,
    textvariable=unit_var,
    values=UQC_OPTIONS,
    state="readonly",
    width=26,
)
unit_combo.grid(row=8, column=1, sticky="ew", padx=(0, 12), pady=6)
make_label(form, "Item taxable value", 8, 2)
item_taxable_entry = make_entry(form, item_taxable_var, 8, 3)

make_label(form, "GST rate %", 9, 0)
rate_entry = make_entry(form, rate_var, 9, 1)
make_label(form, "Tax type", 9, 2)
tax_type_entry = make_entry(form, tax_type_var, 9, 3)
tax_type_entry.config(state="readonly")

add_product_button = tk.Button(
    form,
    text="+ Product Master",
    bg="black",
    fg="white",
    command=add_or_update_product,
)
add_product_button.grid(row=10, column=1, sticky="w", padx=(0, 10), pady=(2, 8))

add_item_button = tk.Button(
    form,
    text="+ Add Item",
    bg="#3f5f8f",
    fg="white",
    command=add_or_update_item,
)
add_item_button.grid(row=10, column=2, sticky="w", padx=(0, 10), pady=(2, 8))

delete_item_button = tk.Button(
    form,
    text="Delete Item",
    bg="#8c1f1f",
    fg="white",
    command=delete_selected_item,
)
delete_item_button.grid(row=10, column=3, sticky="w", padx=(0, 10), pady=(2, 8))

tk.Label(form, text="Tax Calculation", bg="#f2f2f2", font=section_font).grid(
    row=11, column=0, columnspan=4, sticky="w", padx=8, pady=(2, 4)
)

make_label(form, "IGST", 12, 0)
igst_entry = make_entry(form, igst_var, 12, 1)
make_label(form, "CGST", 12, 2)
cgst_entry = make_entry(form, cgst_var, 12, 3)
make_label(form, "SGST", 13, 0)
sgst_entry = make_entry(form, sgst_var, 13, 1)
make_label(form, "Total invoice value", 13, 2)
total_entry = make_entry(form, total_var, 13, 3)

for readonly_entry in [igst_entry, cgst_entry, sgst_entry, total_entry]:
    readonly_entry.config(state="readonly")

invoice_detail_widgets = [
    invoice_date_entry,
    party_gstn_combo,
    supply_type_combo,
    party_name_combo,
    party_state_combo,
    product_combo,
    item_type_combo,
    hsn_entry,
    quantity_entry,
    unit_combo,
    item_taxable_entry,
    rate_entry,
]
invoice_detail_action_widgets = [
    add_party_button,
    add_product_button,
    add_item_button,
    delete_item_button,
]

for variable in [party_gstn_var]:
    variable.trace_add("write", lambda *_args: calculate_tax())

button_bar = tk.Frame(form, bg="#f2f2f2")
button_bar.grid(row=14, column=0, columnspan=4, sticky="ew", padx=8, pady=(10, 4))
for column in range(4):
    button_bar.columnconfigure(column, weight=1)

save_invoice_button = tk.Button(
    button_bar,
    text="+ Add Invoice",
    bg="black",
    fg="white",
    width=14,
    font=("Arial", 9),
    command=save_invoice,
)
save_invoice_button.grid(row=0, column=0, padx=4, pady=4, sticky="ew")

delete_invoice_button = tk.Button(
    button_bar,
    text="Delete Invoice",
    bg="#8c1f1f",
    fg="white",
    width=14,
    font=("Arial", 9),
    state="disabled",
    command=delete_invoice,
)
delete_invoice_button.grid(row=0, column=1, padx=4, pady=4, sticky="ew")

clear_form_button = tk.Button(
    button_bar,
    text="Clear Form",
    bg="#555555",
    fg="white",
    width=14,
    font=("Arial", 9),
    command=lambda: clear_form(keep_header=True),
)
clear_form_button.grid(row=0, column=2, padx=4, pady=4, sticky="ew")

load_draft_button = tk.Button(
    button_bar,
    text="Load Draft",
    bg="#3f5f8f",
    fg="white",
    width=14,
    font=("Arial", 9),
    command=load_draft,
)
load_draft_button.grid(row=0, column=3, padx=4, pady=4, sticky="ew")

import_excel_button = tk.Button(
    button_bar,
    text="Import Excel",
    bg="#3f5f8f",
    fg="white",
    width=14,
    font=("Arial", 9),
    command=import_invoice_excel,
)
import_excel_button.grid(row=1, column=0, padx=4, pady=4, sticky="ew")

export_excel_button = tk.Button(
    button_bar,
    text="Export Excel",
    bg="#3f5f8f",
    fg="white",
    width=14,
    font=("Arial", 9),
    command=export_excel,
)
export_excel_button.grid(row=1, column=1, padx=4, pady=4, sticky="ew")

export_gstr1_button = tk.Button(
    button_bar,
    text="Export GSTR-1 Excel",
    bg="#1f6b3f",
    fg="white",
    width=18,
    font=("Arial", 9),
    command=export_final_gstr1_excel,
)
export_gstr1_button.grid(row=1, column=2, padx=4, pady=4, sticky="ew")

transfer_button = tk.Button(
    button_bar,
    text="Transfer",
    bg="#6b4f1f",
    fg="white",
    width=12,
    font=("Arial", 9),
    command=transfer_to_gstr1_template,
)
transfer_button.grid(row=1, column=3, padx=4, pady=4, sticky="ew")

right_panel = tk.Frame(main, bg="#f2f2f2", width=290)
right_panel.grid(row=1, column=1, sticky="ns", padx=(0, 8), pady=8)
right_panel.grid_propagate(False)
right_panel.rowconfigure(1, weight=1)
right_panel.columnconfigure(0, weight=1)

item_header = tk.Frame(right_panel, bg="#f2f2f2")
item_header.grid(row=0, column=0, sticky="ew", pady=(0, 10))
item_header.columnconfigure(0, weight=1)

tk.Label(
    item_header,
    text="ITEM LIST",
    bg="#f2f2f2",
    font=("Arial", 14, "bold"),
).grid(row=0, column=0, sticky="w")

help_button = tk.Button(
    item_header,
    text="HELP",
    bg="#3f5f8f",
    fg="white",
    width=8,
    font=("Arial", 8, "bold"),
    command=show_help,
)
help_button.grid(row=0, column=1, padx=(6, 4), sticky="e")

about_button = tk.Button(
    item_header,
    text="About",
    bg="#555555",
    fg="white",
    width=8,
    font=("Arial", 8, "bold"),
    command=show_about,
)
about_button.grid(row=0, column=2, sticky="e")

tree_frame = tk.Frame(right_panel, bg="#3f5f8f")
tree_frame.grid(row=1, column=0, sticky="nsew")
tree_frame.rowconfigure(0, weight=1)
tree_frame.columnconfigure(0, weight=1)

item_tree = ttk.Treeview(
    tree_frame,
    columns=("ITEM", "TAXABLE"),
    show="headings",
    selectmode="browse",
)
item_tree.heading("ITEM", text="ITEM NAME")
item_tree.heading("TAXABLE", text="TAXABLE VALUE")
item_tree.column("ITEM", width=180, anchor="w")
item_tree.column("TAXABLE", width=90, anchor="e")
item_tree.grid(row=0, column=0, sticky="nsew")
item_tree.bind("<<TreeviewSelect>>", on_item_select)

scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=item_tree.yview)
scrollbar.grid(row=0, column=1, sticky="ns")
item_tree.configure(yscrollcommand=scrollbar.set)

invoice_strip_panel = tk.Frame(main, bg="#f2f2f2")
invoice_strip_panel.grid(row=2, column=0, columnspan=2, sticky="ew", padx=8, pady=(0, 6))
invoice_strip_panel.columnconfigure(0, weight=1)

tk.Label(
    invoice_strip_panel,
    text="INVOICE LIST",
    bg="#f2f2f2",
    font=("Arial", 12, "bold"),
).grid(row=0, column=0, sticky="w", pady=(0, 6))

invoice_canvas = tk.Canvas(invoice_strip_panel, height=64, bg="#f2f2f2", highlightthickness=0)
invoice_canvas.grid(row=1, column=0, sticky="ew")
invoice_scrollbar = ttk.Scrollbar(invoice_strip_panel, orient="horizontal", command=invoice_canvas.xview)
invoice_scrollbar.grid(row=2, column=0, sticky="ew")
invoice_canvas.configure(xscrollcommand=invoice_scrollbar.set)
invoice_strip_inner = tk.Frame(invoice_canvas, bg="#f2f2f2")
invoice_canvas.create_window((0, 0), window=invoice_strip_inner, anchor="nw")
invoice_strip_inner.bind(
    "<Configure>",
    lambda event: invoice_canvas.configure(scrollregion=invoice_canvas.bbox("all")),
)

status_bar = tk.Label(
    main,
    textvariable=status_var,
    bg="#e5e5e5",
    fg="#222222",
    anchor="w",
    padx=10,
)
status_bar.grid(row=3, column=0, columnspan=2, sticky="ew")

refresh_product_values()
party_gstn_combo["values"] = get_party_gstn_options()
party_name_combo["values"] = get_party_name_options()
update_button_state()
update_quantity_controls()
update_supply_type_controls()
root.after(200, show_introduction)

root.mainloop()
